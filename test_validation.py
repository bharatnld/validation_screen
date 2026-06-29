import streamlit as st
import aiohttp
import asyncio
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_URL = "https://unstandardisable-shantelle-gnawable.ngrok-free.dev/api/process/combinedOcrBytes"

CUSTOMERS = [
    "corybrothers_oil",
    "Cory Brothers Oil CA",   # ← new
]

CA_CUSTOMER_KEY = "corybrothers_oil_ca"   # sent to API

# ── CA schema columns ─────────────────────────────────────────────────────────
# NOTE: header text is the single source of truth for matching against
# CA_NUMERIC_COLS / CA_BADGE_COLS / weight_cols / numeric_map below.
# Keep these header strings in sync everywhere they're referenced.
CA_COLUMNS = [
    ("Billing Document\nVBRK_VBELN",   "shipment",            "root"),
    ("Tied Oil",                       "tied_oil",            "item"),
    ("Excise goods",                   "excise_goods",        "item"),
    ("Load",                           "load",                "item"),
    ("Customs Invoice",                "customs_invoice",     "item"),
    ("Customs Tariff Number",          "hs_code",             "item"),
    ("Country of Origin",              "country_of_origin",   "item"),
    ("Pref. Origin",                   "preferential_origin", "item"),
    ("Net weight\nVBRP_NTGEW",         "net_weight",          "item"),
    ("Gross weight\nVBRP_BRGEW",       "gross_weight",        "item"),
    ("Volume\nVBRP_VOLUM",             "volume_ltrs",         "item"),
    ("Sum of Value",                   "Sum_of_Value",        "item"),
]
# FIX: was 13 entries for 12 columns (off-by-one) — trimmed to match CA_COLUMNS length.
CA_COL_WIDTHS = [16, 10, 12, 10, 16, 14, 20, 10, 10, 14, 14, 14]

# FIX: these sets previously referenced headers that don't exist in CA_COLUMNS
# ("Eaches", "Net Weight", "Total Weight", "Volume (Ltrs)"). Corrected to the
# real header strings (note the literal \n + SAP field-name suffixes).
CA_NUMERIC_COLS = {
    "Net weight\nVBRP_NTGEW",
    "Gross weight\nVBRP_BRGEW",
    "Volume\nVBRP_VOLUM",
    "Sum of Value",
}
CA_BADGE_COLS = {"Tied Oil"}

# Map used by write_subtotal_row to sum numeric columns. FIX: previously used
# header names that don't exist in CA_COLUMNS, so weight/volume subtotals were
# silently always 0. Now keyed by the *actual* CA_COLUMNS headers.
CA_SUBTOTAL_NUMERIC_MAP = {
    "Net weight\nVBRP_NTGEW":   "net_weight",
    "Gross weight\nVBRP_BRGEW": "gross_weight",
    "Volume\nVBRP_VOLUM":       "volume_ltrs",
    "Sum of Value":             "Sum_of_Value",
}

# ── SAP schema columns ────────────────────────────────────────────────────────
SAP_COLUMNS = [
    ("Billing Document\nVBRK_VBELN",                                    "invoice_number",       "inv"),
    ("Due Date",                                                         "due_date",             "inv"),
    ("Document number of the\nreference document\nVBRP_VGBEL",          "delivery_note_number", "item"),
    ("Material Number",                                                  "material_number",      "item"),
    ("Material Name\nMAKT_MAKTX",                                       "description",          "item"),
    ("Billing quantity in\nstockkeeping unit\nVBRP_FKLMG",              "quantity",             "item"),
    ("UoM",                                                              "quantity_uom",         "item"),
    ("Net weight\nVBRP_NTGEW",                                          "nett_weight",          "item"),
    ("Gross weight\nVBRP_BRGEW",                                        "gross_weight",         "item"),
    ("Volume\nVBRP_VOLUM",                                              "volume",               "item"),
    ("Volume unit\nVBRP_VOLEH",                                         "volume_uom",           "item"),
    ("Net value of the billing item\nin document currency\nVBRP_NETWR", "line_net_price",       "item"),
    ("Rebate basis 1\nVBRP_BONBA",                                      "rebate_basis_amount",  "item"),
    ("SD Document Currency\nVBRK_WAERK",                                "currency",             "inv"),
    ("Country of Origin",                                               "country_of_origin",    "item"),
    ("Customs Tariff Number",                                           "hs_code",              "item"),
    ("Pref. Origin",                                                    "preferential_origin",  "item"),
    ("IncoTerms",                                                       "incoterms",            "item"),
    ("Unit Price",                                                      "unit_price",           "item"),
    ("Line Amount",                                                     "line_amount",          "item"),
]
SAP_UI_LABELS = [
    "Billing Doc", "Due Date", "Del. Note No.", "Material No.", "Material Name",
    "Qty", "UoM", "Net Weight", "Gross Weight", "Volume", "Vol. Unit",
    "Net Value", "Rebate Basis", "Currency", "Country of Origin",
    "Customs Tariff No.", "Pref. Origin", "IncoTerms", "Unit Price", "Line Amount",
]
SAP_COL_WIDTHS = [20, 14, 24, 16, 40, 10, 8, 12, 12, 10, 10, 18, 14, 12, 14, 18, 12, 18, 12, 14]
SAP_NUMERIC_COLS = {"Net Value", "Unit Price", "Line Amount", "Rebate Basis"}
SAP_WEIGHT_COLS  = {"Net Weight", "Gross Weight", "Volume", "Qty"}
SAP_BADGE_COL    = "Pref. Origin"

EU_COUNTRIES = {"BE","DE","FR","IT","NL","PL","ES","PT","AT","SE","FI","DK",
                "IE","CZ","SK","HU","RO","BG","HR","SI","EE","LV","LT","LU","MT","CY","GR"}


def _is_non_eu_code(coo) -> bool:
    """
    FIX: original code called coo.upper() without guarding against None,
    which raises AttributeError whenever country_of_origin is missing/null
    from OCR extraction (a realistic case). Centralized + null-safe here.
    """
    coo_str = "" if coo is None else str(coo).strip()
    return len(coo_str) == 2 and coo_str.upper() not in EU_COUNTRIES


# ── API ───────────────────────────────────────────────────────────────────────

async def call_api(files_data: list, customer_name: str):
    try:
        data = aiohttp.FormData()
        for file_bytes, file_name in files_data:
            data.add_field("files", file_bytes, filename=file_name, content_type="application/pdf")
        data.add_field("customer_name", customer_name)
        data.add_field("model_name", "gpt-4o")
        async with aiohttp.ClientSession() as session:
            async with session.post(API_URL, data=data, timeout=aiohttp.ClientTimeout(total=300)) as response:
                return await response.json()
    except Exception as e:
        return {"error": str(e)}


# ── SAP helpers ───────────────────────────────────────────────────────────────

def merge_delivery_into_invoices(invoices: list) -> list:
    lookup: dict = {}
    for inv in invoices:
        for item in inv.get("line_items", []):
            dn  = item.get("delivery_note_number")
            seq = item.get("item_seq")
            if dn and seq:
                key = (str(dn), str(seq))
                if key not in lookup:
                    lookup[key] = item
    DR_FILL = ["material_number", "volume", "volume_uom", "nett_weight", "gross_weight",
               "hs_code", "country_of_origin", "preferential_origin", "incoterms", "quantity_uom"]
    for inv in invoices:
        for item in inv.get("line_items", []):
            dn  = item.get("delivery_note_number")
            seq = item.get("item_seq")
            if not dn or not seq:
                continue
            src = lookup.get((str(dn), str(seq)), {})
            for f in DR_FILL:
                if not item.get(f) and src.get(f):
                    item[f] = src[f]
    return invoices


def build_sap_excel(invoices: list) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Invoice Data"
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="EEF2F7")
    inv_fill = PatternFill("solid", fgColor="D6E4F0")

    for ci, (hdr, _, _) in enumerate(SAP_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 52

    row = 2
    for inv in invoices:
        for i, item in enumerate(inv.get("line_items", [])):
            rf = inv_fill if i == 0 else (alt_fill if row % 2 == 0 else None)
            for ci, (_, field, src) in enumerate(SAP_COLUMNS, 1):
                val = inv.get(field, "") if src == "inv" else item.get(field, "")
                if val is None:
                    val = ""
                c = ws.cell(row=row, column=ci, value=val)
                c.font = Font(name="Arial", size=9); c.alignment = left; c.border = border
                if rf:
                    c.fill = rf
            row += 1

    for i, w in enumerate(SAP_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SAP_COLUMNS))}1"
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()


# ── CA helpers ────────────────────────────────────────────────────────────────

def parse_ca_response(result: dict):
    """Extract CA data: returns (shipment_str, items_list)"""
    try:
        ca = result["data"]["extracted_data"]["gpt_extraction_output"]
        shipment = ca.get("shipment", "")
        items    = ca.get("items", [])
        for it in items:
            it.setdefault("shipment", shipment)
        return shipment, items
    except (KeyError, TypeError):
        return "", []


def build_ca_excel(shipment: str, items: list) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "CA Report"

    hdr_fill = PatternFill("solid", fgColor="1A3C5E")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    sub_fill = PatternFill("solid", fgColor="2E4D3A")   # dark green subtotals
    sub_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    tot_fill = PatternFill("solid", fgColor="3B5225")
    tot_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center")
    right    = Alignment(horizontal="right",  vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    yes_fill = PatternFill("solid", fgColor="FFF3CD")    # amber for tied/excise YES
    non_eu_fill = PatternFill("solid", fgColor="FFE4E4") # light red for non-EU

    # Header
    headers = [c[0] for c in CA_COLUMNS]
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 28

    # Separate N/A and YES tied oil groups
    na_items  = [it for it in items if str(it.get("tied_oil", "")).upper() != "YES"]
    yes_items = [it for it in items if str(it.get("tied_oil", "")).upper() == "YES"]

    def write_item_row(ws_row, item, row_fill=None):
        coo = item.get("country_of_origin", "")
        coo_str = "" if coo is None else str(coo)
        is_non_eu = _is_non_eu_code(coo_str)  # FIX: null-safe now

        for ci, (hdr, field, src) in enumerate(CA_COLUMNS, 1):
            if src == "root":
                val = shipment
            else:
                val = item.get(field, "")
            if val is None:
                val = ""

            c = ws.cell(row=ws_row, column=ci, value=val)
            c.font   = Font(name="Arial", size=9)
            c.border = border

            if hdr == "Country of Origin" and is_non_eu and val not in ("", "EU", "EU not preferential"):
                c.fill = non_eu_fill
                c.font = Font(name="Arial", size=9, color="CC0000", bold=True)
                c.alignment = left
            elif hdr in CA_BADGE_COLS and str(val).upper() == "YES":
                c.fill = yes_fill
                c.font = Font(name="Arial", size=9, bold=True, color="7D4A00")
                c.alignment = center
            elif hdr in CA_NUMERIC_COLS:
                try:
                    c.value = float(val) if val != "" else ""
                    c.number_format = '#,##0.000' if hdr != "Sum of Value" else '#,##0.00'
                    c.alignment = right
                except (ValueError, TypeError):
                    c.alignment = right
            else:
                c.alignment = left
                if row_fill:
                    c.fill = row_fill

    def write_subtotal_row(ws_row, label, group_items, fill, font):
        ws.cell(row=ws_row, column=1, value=label).font = font
        ws.cell(row=ws_row, column=1).fill = fill
        ws.cell(row=ws_row, column=1).border = border
        ws.cell(row=ws_row, column=1).alignment = left

        for ci in range(2, len(CA_COLUMNS) + 1):
            c = ws.cell(row=ws_row, column=ci)
            c.fill = fill; c.border = border

        # FIX: use CA_SUBTOTAL_NUMERIC_MAP (correct headers) instead of the old
        # numeric_map that referenced nonexistent headers like "Eaches" / "Total Weight".
        for hdr, field, _ in CA_COLUMNS:
            if hdr in CA_SUBTOTAL_NUMERIC_MAP:
                map_field = CA_SUBTOTAL_NUMERIC_MAP[hdr]
                ci = [c[0] for c in CA_COLUMNS].index(hdr) + 1
                total = sum(float(it.get(map_field, 0) or 0) for it in group_items)
                c = ws.cell(row=ws_row, column=ci, value=total)
                c.font = font; c.fill = fill; c.border = border
                c.number_format = '#,##0.000' if hdr != "Sum of Value" else '#,##0.00'
                c.alignment = right

    row = 2

    # N/A group
    for item in na_items:
        rf = alt_fill if row % 2 == 0 else None
        write_item_row(row, item, rf)
        row += 1
    write_subtotal_row(row, "Totaal N/A", na_items, sub_fill, sub_font)
    row += 1

    # YES group
    for item in yes_items:
        rf = alt_fill if row % 2 == 0 else None
        write_item_row(row, item, rf)
        row += 1
    write_subtotal_row(row, "Totaal YES", yes_items, sub_fill, sub_font)
    row += 1

    # Grand total
    write_subtotal_row(row, f"Totaal {shipment}", items, tot_fill, tot_font)
    row += 1
    write_subtotal_row(row, "Eindtotaal", items, tot_fill, tot_font)

    for i, w in enumerate(CA_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()


# ── HTML table builder (shared) ───────────────────────────────────────────────

def render_html_table(headers: list, rows_data: list, badge_cols: set,
                       numeric_cols: set, weight_cols: set = None,
                       first_col_sticky=True) -> str:
    th_cells = "".join(
        f'<th style="{"position:sticky;left:0;z-index:3;" if i == 0 and first_col_sticky else ""}">'
        f'{h.replace(chr(10), "<br>")}</th>'
        for i, h in enumerate(headers)
    )
    tbody = ""
    for row_vals, row_class in rows_data:
        tds = ""
        coo_val = ""
        for h, v in zip(headers, row_vals):
            if h == "Country of Origin":
                coo_val = "" if v is None else str(v)

        for h, v in zip(headers, row_vals):
            v = "" if v is None else v
            cell_html = ""

            if h in badge_cols:
                sv = str(v).upper()
                if sv == "YES":
                    cell_html = '<span class="badge-yes">YES</span>'
                elif sv == "NO":
                    cell_html = '<span class="badge-no">NO</span>'
                elif sv == "N/A":
                    cell_html = '<span class="badge-na">N/A</span>'
                else:
                    cell_html = str(v)
            elif h == "Country of Origin":
                is_non_eu = _is_non_eu_code(coo_val)  # FIX: null-safe now
                if is_non_eu and coo_val not in ("", "EU", "EU not preferential"):
                    cell_html = f'<span class="badge-non-eu">{v}</span>'
                else:
                    cell_html = str(v)
            elif h == "Pref. Origin":
                sv = str(v).upper()
                if sv == "YES":
                    cell_html = '<span class="badge-yes">YES</span>'
                elif sv == "NO":
                    cell_html = '<span class="badge-no">NO</span>'
                else:
                    cell_html = str(v)
            elif h in numeric_cols:
                try:
                    cell_html = f"{float(v):,.2f}" if v != "" else ""
                except (ValueError, TypeError):
                    cell_html = str(v)
            elif weight_cols and h in weight_cols:
                try:
                    cell_html = f"{float(v):,.3f}" if v != "" else ""
                except (ValueError, TypeError):
                    cell_html = str(v)
            else:
                cell_html = str(v)

            tds += f"<td>{cell_html}</td>"
        tbody += f"<tr class='{row_class}'>{tds}</tr>"

    return f"""
    <div class="excel-wrap">
      <div class="excel-scroll">
        <table class="excel-table">
          <thead><tr>{th_cells}</tr></thead>
          <tbody>{tbody}</tbody>
        </table>
      </div>
    </div>"""


# ── PAGE CONFIG ───────────────────────────────────────────────────────────────

st.set_page_config(page_title="Invoice Extractor", page_icon="📄", layout="wide")

st.markdown("""
<style>
    .stApp { background: #F7F9FC; }
    .main-title { font-size:26px; font-weight:700; color:#1F3864; margin-bottom:2px; }
    .sub-title  { font-size:14px; color:#6B7280; margin-bottom:20px; }

    .stat-card  { background:white; border-radius:10px; padding:14px 18px;
                  border:1px solid #E5E7EB; text-align:center; }
    .stat-num   { font-size:26px; font-weight:700; color:#1F3864; }
    .stat-label { font-size:11px; color:#6B7280; margin-top:2px; }

    .file-pill     { display:inline-block; background:#EEF2F7; color:#1F3864;
                     border-radius:20px; padding:3px 12px; font-size:12px;
                     margin:3px 4px 3px 0; font-weight:500; }
    .file-pill.dr  { background:#FEF3C7; color:#92400E; }

    div[data-testid="stFileUploader"] { background:white; border-radius:10px;
        border:1.5px dashed #CBD5E1; padding:12px; }

    .excel-wrap   { border-radius:10px; overflow:hidden; border:1px solid #D1D5DB;
                    box-shadow:0 2px 8px rgba(0,0,0,0.06); }
    .excel-scroll { overflow-x:auto; overflow-y:auto; max-height:520px; }
    .excel-table  { border-collapse:collapse; font-family:Arial,sans-serif;
                    font-size:12px; white-space:nowrap; min-width:100%; }
    .excel-table thead th { background:#1F3864; color:white; padding:8px 12px;
        text-align:center; font-weight:600; font-size:11px;
        position:sticky; top:0; z-index:2;
        border-right:1px solid #2d4d7a; border-bottom:2px solid #152849; }
    .excel-table thead th:first-child { position:sticky; left:0; z-index:3; }
    .excel-table tbody td { padding:6px 12px; border-right:1px solid #E5E7EB;
                             border-bottom:1px solid #E5E7EB; color:#111827; }
    .excel-table tbody td:first-child { position:sticky; left:0; z-index:1;
        font-weight:600; color:#1F3864; background:inherit; }
    .excel-table tbody tr:nth-child(odd)  { background:#FFFFFF; }
    .excel-table tbody tr:nth-child(even) { background:#EEF2F7; }
    .excel-table tbody tr.inv-first { background:#D6E4F0 !important; }
    .excel-table tbody tr:hover td  { background:#DBEAFE !important; }

    .badge-yes    { display:inline-block; background:#D1FAE5; color:#065F46;
                    border-radius:10px; padding:1px 8px; font-size:11px; font-weight:600; }
    .badge-no     { display:inline-block; background:#FEE2E2; color:#991B1B;
                    border-radius:10px; padding:1px 8px; font-size:11px; font-weight:600; }
    .badge-na     { display:inline-block; background:#F3F4F6; color:#6B7280;
                    border-radius:10px; padding:1px 8px; font-size:11px; font-weight:600; }
    .badge-non-eu { display:inline-block; background:#FEE2E2; color:#CC0000;
                    border-radius:10px; padding:1px 8px; font-size:11px; font-weight:700; }

    .section-title { font-size:15px; font-weight:600; color:#1F3864;
                     margin:20px 0 10px 0; }
    .row-count     { font-size:12px; font-weight:400; color:#6B7280;
                     background:#F3F4F6; border-radius:20px; padding:1px 10px; }
    .customer-tag  { display:inline-block; background:#1A3C5E; color:white;
                     border-radius:20px; padding:3px 14px; font-size:12px;
                     font-weight:600; margin-left:10px; vertical-align:middle; }
</style>
""", unsafe_allow_html=True)

# ── HEADER ────────────────────────────────────────────────────────────────────

st.markdown('<div class="main-title">📄 Invoice Data Extractor</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Upload invoices and delivery receipts — data is merged and exported to Excel.</div>', unsafe_allow_html=True)

col_up, col_cust = st.columns([3, 2])
with col_up:
    uploaded_files = st.file_uploader(
        "Upload PDFs", type=["pdf"],
        accept_multiple_files=True, label_visibility="collapsed",
    )
with col_cust:
    customer_display = st.selectbox("Customer", CUSTOMERS, index=0)

is_ca = (customer_display == "Cory Brothers Oil CA")
api_customer = CA_CUSTOMER_KEY if is_ca else customer_display

if uploaded_files:
    pills_html = ""
    for f in uploaded_files:
        is_dr = any(k in f.name.lower() for k in ["delivery", "receipt", "dr", "dn", "despatch"])
        cls  = "file-pill dr" if is_dr else "file-pill"
        icon = "📦" if is_dr else "🧾"
        pills_html += f'<span class="{cls}">{icon} {f.name}</span>'
    st.markdown(pills_html, unsafe_allow_html=True)

    st.divider()
    if st.button("⚡ Extract & Merge Data", use_container_width=True, type="primary"):
        with st.spinner(f"Processing {len(uploaded_files)} file(s)..."):
            files_data = [(f.getvalue(), f.name) for f in uploaded_files]
            result = asyncio.run(call_api(files_data, api_customer))

        if "error" in result:
            st.error(f"API error: {result['error']}")
            st.stop()

        if is_ca:
            shipment, ca_items = parse_ca_response(result)
            if not ca_items:
                st.error("No CA items found in API response.")
                with st.expander("Raw response"):
                    st.json(result)
                st.stop()
            st.session_state["mode"]     = "ca"
            st.session_state["shipment"] = shipment
            st.session_state["ca_items"] = ca_items
        else:
            try:
                invoices = result["data"]["extracted_data"]["gpt_extraction_output"]["invoices"]
            except (KeyError, TypeError):
                st.error("Unexpected response format.")
                with st.expander("Raw response"):
                    st.json(result)
                st.stop()
            invoices = merge_delivery_into_invoices(invoices)
            st.session_state["mode"]     = "sap"
            st.session_state["invoices"] = invoices

        st.session_state["uploaded_names"] = [f.name for f in uploaded_files]
        st.session_state["n_files"]        = len(uploaded_files)
        st.session_state["customer"]       = customer_display

# ══════════════════════════════════════════════════════════════════════════════
# CA MODE
# ══════════════════════════════════════════════════════════════════════════════

if st.session_state.get("mode") == "ca":
    shipment = st.session_state["shipment"]
    ca_items = st.session_state["ca_items"]
    n_files  = st.session_state.get("n_files", 0)

    # FIX: "eaches" was computed but never used/displayed anywhere — removed.
    # FIX: Sum_of_Value total now guarded against None/garbage values consistently.
    total_value = sum(float(it.get("Sum_of_Value", 0) or 0) for it in ca_items)

    st.markdown('### Summary <span class="customer-tag">🚢 Cory Brothers Oil CA</span>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="stat-card"><div class="stat-num">{n_files}</div><div class="stat-label">Files uploaded</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-card"><div class="stat-num">{shipment or "—"}</div><div class="stat-label">Shipment</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-card"><div class="stat-num">{len(ca_items)}</div><div class="stat-label">Line items</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="stat-card"><div class="stat-num">€{total_value:,.2f}</div><div class="stat-label">Total value</div></div>', unsafe_allow_html=True)

    st.markdown(f'<div class="section-title">📊 CA Report Preview <span class="row-count">{len(ca_items)} rows · {len(CA_COLUMNS)} columns</span></div>', unsafe_allow_html=True)

    headers = [c[0] for c in CA_COLUMNS]
    rows_data = []
    for i, item in enumerate(ca_items):
        row_vals = []
        for hdr, field, src in CA_COLUMNS:
            val = shipment if src == "root" else item.get(field, "")
            row_vals.append("" if val is None else val)
        rows_data.append((row_vals, "inv-first" if i == 0 else ""))

    # FIX: weight_cols now passes the real CA_COLUMNS header strings (with \n +
    # SAP suffixes) instead of nonexistent labels like "Eaches"/"Total Weight".
    st.markdown(render_html_table(
        headers, rows_data,
        badge_cols=CA_BADGE_COLS,
        numeric_cols={"Sum of Value"},
        weight_cols={"Net weight\nVBRP_NTGEW", "Gross weight\nVBRP_BRGEW", "Volume\nVBRP_VOLUM"},
    ), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    fname = st.session_state.get("uploaded_names", ["ca_report"])[0].replace(".pdf", "")
    excel_bytes = build_ca_excel(shipment, ca_items)
    st.download_button(
        label="⬇️ Download CA Report Excel",
        data=excel_bytes,
        file_name=f"{fname}_ca_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

# ══════════════════════════════════════════════════════════════════════════════
# SAP MODE
# ══════════════════════════════════════════════════════════════════════════════

elif st.session_state.get("mode") == "sap":
    invoices = st.session_state["invoices"]
    n_files  = st.session_state.get("n_files", 0)
    total_items = sum(len(inv.get("line_items", [])) for inv in invoices)
    total_value = sum(float(inv.get("net_price_total", 0) or 0) for inv in invoices)

    cust_label = st.session_state.get("customer", "")
    st.markdown(f'### Summary <span class="customer-tag">🧾 {cust_label}</span>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="stat-card"><div class="stat-num">{n_files}</div><div class="stat-label">Files uploaded</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-card"><div class="stat-num">{len(invoices)}</div><div class="stat-label">Invoices found</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-card"><div class="stat-num">{total_items}</div><div class="stat-label">Line items</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="stat-card"><div class="stat-num">€{total_value:,.2f}</div><div class="stat-label">Total net value</div></div>', unsafe_allow_html=True)

    st.markdown(f'<div class="section-title">📊 Data preview <span class="row-count">{total_items} rows · {len(SAP_COLUMNS)} columns</span></div>', unsafe_allow_html=True)

    headers   = SAP_UI_LABELS
    rows_data = []
    for inv in invoices:
        for i, item in enumerate(inv.get("line_items", [])):
            row_vals = []
            for lbl, (_, field, src) in zip(SAP_UI_LABELS, SAP_COLUMNS):
                val = inv.get(field, "") if src == "inv" else item.get(field, "")
                row_vals.append("" if val is None else val)
            rows_data.append((row_vals, "inv-first" if i == 0 else ""))

    st.markdown(render_html_table(
        headers, rows_data,
        badge_cols=set(),
        numeric_cols=SAP_NUMERIC_COLS,
        weight_cols=SAP_WEIGHT_COLS,
    ), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col_filter, col_dl = st.columns([3, 2])
    with col_filter:
        inv_nums = ["All invoices"] + [inv.get("invoice_number", "?") for inv in invoices]
        selected = st.selectbox("Filter by invoice", inv_nums, label_visibility="collapsed")
    with col_dl:
        export_inv = invoices if selected == "All invoices" else \
                     [inv for inv in invoices if inv.get("invoice_number") == selected]
        fname = st.session_state.get("uploaded_names", ["invoices"])[0].replace(".pdf", "")
        if selected != "All invoices":
            fname = f"invoice_{selected}"
        st.download_button(
            label="⬇️ Download Excel (SAP format)",
            data=build_sap_excel(export_inv),
            file_name=f"{fname}_extracted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary",
        )

    if selected != "All invoices":
        filtered = [inv for inv in invoices if inv.get("invoice_number") == selected]
        fi = sum(len(inv.get("line_items", [])) for inv in filtered)
        st.caption(f"Showing {fi} line item(s) for invoice {selected}")
        frows = []
        for inv in filtered:
            for i, item in enumerate(inv.get("line_items", [])):
                rv = []
                for lbl, (_, field, src) in zip(SAP_UI_LABELS, SAP_COLUMNS):
                    val = inv.get(field, "") if src == "inv" else item.get(field, "")
                    rv.append("" if val is None else val)
                frows.append((rv, "inv-first" if i == 0 else ""))
        st.markdown(render_html_table(
            headers, frows,
            badge_cols=set(), numeric_cols=SAP_NUMERIC_COLS, weight_cols=SAP_WEIGHT_COLS,
        ), unsafe_allow_html=True)
